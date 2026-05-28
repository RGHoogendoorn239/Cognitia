# Knowledge Claims Tool – CIMO, Toulmin, Denyer, Aristotelos
# Voldoet aan eisen in Knowledge claims folder
# By Raymond Hoogendoorn, Copyright 2026

"""
Module for the three Knowledge-claims tools based on the documents in the Knowledge claims folder:
- Toulmin/CIMO analysis: analyse and qualify CIMO conjectures
- Denyer: structure semantic networks according to CIMO logic
- Aristoteles: build a semantic network around a central concept
"""

from __future__ import annotations

import io
import json
import re
from typing import Any, Callable, List, Optional, Tuple

# Excel-export: gebruik openpyxl indien beschikbaar, anders CSV
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# --- Kennisclaims-hierarchie (2025 11 13 kennisclaims_hierarchie) ---
# Exacte 10 niveaus uit PDF/Excel; als xlsx wordt meegegeven wordt die gebruikt, anders deze definitie.

HIERARCHY_LEVELS = [
    # 1. Common sense
    {"nummer": 1, "niveau": "Common sense", "zekerheid": "Very low", "explicietheid": "Implicit",
     "omschrijving": "Everyday intuitive knowledge, often unfounded",
     "indicatoren": "Spontaneous statements, experiential sayings, no explicit justification",
     "procedure": "Observation of language use and absence of testing"},
    # 2. Assumption
    {"nummer": 2, "niveau": "Assumption", "zekerheid": "Low", "explicietheid": "Implicit",
     "omschrijving": "Taken-for-granted starting point without evidence",
     "indicatoren": "Unquestioned assumptions, often tacit",
     "procedure": "Analysis of reasoning and detection of unsupported assumptions"},
    # 3. Idea
    {"nummer": 3, "niveau": "Idea", "zekerheid": "Low", "explicietheid": "Exploratory / semi-explicit",
     "omschrijving": "First thought or concept, not yet tested",
     "indicatoren": "Creative formulation, suggestive wording, open formulation",
     "procedure": "Identification of exploratory statements without empirical testing"},
    # 4. Hypothesis
    {"nummer": 4, "niveau": "Hypothesis", "zekerheid": "Medium", "explicietheid": "Explicit",
     "omschrijving": "Testable assumption based on observation",
     "indicatoren": "Formulated as \"if… then…\", based on observation or literature",
     "procedure": "Scientific testing via experiment, survey or simulation"},
    # 5. Conjecture (practice-based)
    {"nummer": 5, "niveau": "Conjecture (practice-based)", "zekerheid": "Medium", "explicietheid": "Contextually explicit",
     "omschrijving": "Practice-oriented hypothesis based on experience",
     "indicatoren": "\"If… then…\" statements in context, often in CIMO form",
     "procedure": "Reflection on practice experience, triangulation with observation and feedback"},
    # 6. Experiential knowledge
    {"nummer": 6, "niveau": "Experiential knowledge", "zekerheid": "Fair", "explicietheid": "Implicit",
     "omschrijving": "Knowledge from repeated application in practice",
     "indicatoren": "Repetition of actions, implicit routines, narrative descriptions",
     "procedure": "Analysis of practice stories, observation of behaviour patterns"},
    # 7. Application knowledge
    {"nummer": 7, "niveau": "Application knowledge", "zekerheid": "High in context", "explicietheid": "Practice-oriented explicit",
     "omschrijving": "Knowledge that proves effective in concrete situations",
     "indicatoren": "Successful application, repeatability in context, user feedback",
     "procedure": "Evaluation of interventions in context, validation via practice results"},
    # 8. Expertise
    {"nummer": 8, "niveau": "Expertise", "zekerheid": "High", "explicietheid": "Domain-specific explicit",
     "omschrijving": "Validated knowledge of experts within a domain",
     "indicatoren": "Recognition by peers, publications, certification, consistent application",
     "procedure": "Peer review, portfolio analysis, domain-specific assessment"},
    # 9. Scientific knowledge
    {"nummer": 9, "niveau": "Scientific knowledge", "zekerheid": "Very high", "explicietheid": "Generally explicit",
     "omschrijving": "Systematically tested and replicated knowledge",
     "indicatoren": "Publications, replications, statistical validity, theoretical justification",
     "procedure": "Empirical research, systematic review, meta-analysis"},
    # 10. Theory
    {"nummer": 10, "niveau": "Theory", "zekerheid": "Highest", "explicietheid": "Fully explicit",
     "omschrijving": "Validated explanatory model with broad applicability",
     "indicatoren": "Conceptual models, explanatory power, cross-domain application",
     "procedure": "Theoretical testing, model validation, application in diverse contexts"},
]

# Text version for the system prompt (table as in PDF/Excel)
HIERARCHY_LEVELS_AND_INDICATORS = """
## Hierarchy of knowledge levels (2025 11 13 kennisclaims_hierarchie)

Use **only** the table below to determine the knowledge level. For each conjecture: (1) search in the documents and in the Cognitia passages for **indicators** and **procedure references** that match a level; (2) choose the **number (1–10)** and **level** that fit best; (3) also report **explicitness** and **certainty level** as in the table.

| Number | Level | Description | Indicators | Procedure | Certainty level | Explicitness |
|--------|-------|-------------|------------|-----------|-----------------|--------------|
| 1 | Common sense | Everyday intuitive knowledge, often unfounded | Spontaneous statements, experiential sayings, no explicit justification | Observation of language use and absence of testing | Very low | Implicit |
| 2 | Assumption | Taken-for-granted starting point without evidence | Unquestioned assumptions, often tacit | Analysis of reasoning and detection of unsupported assumptions | Low | Implicit |
| 3 | Idea | First thought or concept, not yet tested | Creative formulation, suggestive wording, open formulation | Identification of exploratory statements without empirical testing | Low | Exploratory / semi-explicit |
| 4 | Hypothesis | Testable assumption based on observation | Formulated as \"if… then…\", based on observation or literature | Scientific testing via experiment, survey or simulation | Medium | Explicit |
| 5 | Conjecture (practice-based) | Practice-oriented hypothesis based on experience | Contextualised \"if… then…\" statements, often in CIMO form | Reflection on practice experience, triangulation with observation and feedback | Medium | Contextually explicit |
| 6 | Experiential knowledge | Knowledge from repeated application in practice | Repetition of actions, implicit routines, narrative descriptions | Analysis of practice stories, observation of behaviour patterns | Fair | Implicit |
| 7 | Application knowledge | Knowledge that proves effective in concrete situations | Successful application, repeatability in context, user feedback | Evaluation of interventions in context, validation via practice results | High in context | Practice-oriented explicit |
| 8 | Expertise | Validated knowledge of experts within a domain | Recognition by peers, publications, certification, consistent application | Peer review, portfolio analysis, domain-specific assessment | High | Domain-specific explicit |
| 9 | Scientific knowledge | Systematically tested and replicated knowledge | Publications, replications, statistical validity, theoretical justification | Empirical research, systematic review, meta-analysis | Very high | Generally explicit |
| 10 | Theory | Validated explanatory model with broad applicability | Conceptual models, explanatory power, cross-domain application | Theoretical testing, model validation, application in diverse contexts | Highest | Fully explicit |
"""

# --- Systeemprompts uit Knowledge claims documenten ---

def _build_toulmin_system_prompt(hierarchy_text: Optional[str] = None) -> str:
    hierarchy_block = hierarchy_text.strip() if hierarchy_text else HIERARCHY_LEVELS_AND_INDICATORS
    return f"""Communicate only in English.
You are the **GDGO-Toulmin-Weaver** agent: you translate CIMO conjectures into Toulmin-structured knowledge claims and determine the knowledge level (1–10) based on the official knowledge-claims hierarchy. You work in a four-phase architecture: you receive CIMO claims, optional evidence documents, and optionally relevant passages from the Cognitia document base.

Very important:
- When the user message contains a block titled **"Relevante kennis uit de Cognitia documentenbase (literatuur)"**, those are already the relevant passages from Cognitia.
- You MUST use those passages directly to assess certainty/knowledge level.
- You MUST NOT ask the user again to provide or paste passages from Cognitia. They are already provided in the conversation.
For each conjecture: (1) recognise CIMO, (2) search in all provided texts (CIMO document, evidence documents, and the Cognitia block) for **indicators** and **procedure references** from the hierarchy table, (3) map these findings to exactly one level (number 1–10), (4) fill the Toulmin structure (claim, data, warrant, qualifier).

{hierarchy_block}

## Input
You receive:
1. A document with (clustered) CIMO conjectures.
2. Optionally: document(s) with data or evidence as support.

## Procedure

### 1. Recognising the CIMO structure
For each conjecture, identify:
- **Context (C)**: Where does it take place?
- **Intervention (I)**: What is being applied?
- **Mechanism (M)**: What is activated?
- **Outcome (O)**: What is the result?

### 2. Searching for indicators and procedures (required for knowledge level)
For **each** conjecture, search the provided documents and explicitly note:
- **Found indicators**: Is there an \"if… then…\" structure? Reflection on practice experience? Mention of repeatability, validation or theoretical underpinning? References?
- **Procedure references**: Is methodology, research design, validation steps or testing described?

### 3. Determining knowledge level and certainty (based on step 2 + optionally Cognitia)
- Map the **found** indicators and procedure references to the hierarchy above.
- **If there is a block \"Relevante kennis uit de Cognitia documentenbase\" (or similar):** use those passages to determine the certainty/knowledge level. Check whether the conjectures/claims are supported, validated or theoretically underpinned in the literature (Cognitia). This can raise the level (e.g. from conjecture to application knowledge or expertise) or refine the qualifier.
- Determine the knowledge level of each conjecture and **explicitly justify**: which indicators/procedures (and, if applicable, which Cognitia passages) lead to this level? If you do not find clear indicators, choose a lower level and motivate this.

### 4. Toulmin structure
For each conjecture, provide:
- **Claim**: The central statement.
- **Data**: Supporting data or observations (with reference if possible).
- **Warrant**: The reason why the data support the claim (with reference).
- **Qualifier**: The context or limitations of validity (with reference).

## Output (GDGO-Toulmin-Weaver)
Present each conjecture in the following fixed order. Use the exact terms from the hierarchy table:
1. **CIMO analysis**: explicitly write  
   - `Context (C): ...`  
   - `Intervention (I): ...`  
   - `Mechanism (M): ...`  
   - `Outcome (O): ...`
2. **Found indicators**: start the line with `Found indicators:` and give the indicators (verbatim from documents and, where applicable, Cognitia passages).
3. **Procedure references**: start the line with `Procedure references:` and list the procedures (if present).
4. **Level_number**: start the line with `Level_number:` and give the number 1–10 (from the table).
5. **Level_label**: start the line with `Level_label:` and give the exact name of the level (e.g. "Conjecture (practice-based)", "Application knowledge").
6. **Certainty level**: start the line with `Certainty level:` and use the wording from the table (Very low, Low, Medium, Fair, High in context, High, Very high, Highest).
7. **Explicitness**: start the line with `Explicitness:` and use the wording from the table (Implicit, Contextually explicit, etc.).
8. **Justification**: start the line with `Justification:` and explain which indicators/procedures (and, if applicable, which Cognitia passages) lead to this level.
9. **Toulmin**: give the Toulmin structure with **four separate lines**, starting exactly as:
   - `Claim:` the central statement  
   - `Data:` the supporting data or observations (with reference if possible)  
   - `Warrant:` the reason why the data support the claim (with reference)  
   - `Qualifier:` the context or limitations of validity (with reference)

Give this full block (steps 1–9) for **each conjecture**. Be explicit and structured. In case of doubt, choose the lower level (lower number) and justify."""

DENYER_CIMO_SYSTEM_PROMPT = """Communicate only in English.
You are a Copilot agent specialised in analysing semantic networks and structuring practice knowledge according to the CIMO logic. You support researchers, policy advisors and educational designers in making conjectures explicit and developing knowledge claims.

## CIMO logic
- **Context (C)**: The situation or environment in which an intervention takes place.
- **Intervention (I)**: The action, measure or approach that is applied.
- **Mechanism (M)**: The underlying mechanism that is activated by the intervention.
- **Outcome (O)**: The result or effect that follows from the mechanism in that context.

Formulate knowledge claims as conjectures:
\"In [Context], when [Intervention] is applied, this activates [Mechanism], which leads to [Outcome].\"

## Procedure
1. **Extraction**: Read the semantic network. Identify central concepts, related concepts, relations between concepts, and influences on those relations.
2. **Conceptual CIMO rows**: For each cluster of related concepts, construct a conceptual CIMO row with domain-specific concepts.
3. **Conjectures**: For each CIMO row, formulate a hypothesis in natural language. Explicitly label the status as \"conjecture\". Use active formulations.
4. **Tables**: For each conceptual structure, generate a table with rows in which each conjecture is represented in CIMO form (columns: Context, Intervention, Mechanism, Outcome).

## Behavioural guidelines
- Be analytical, precise and transparent.
- Always state the status of knowledge (conjecture, hypothesis, claim).
- Avoid speculation beyond the network.
- Respect domain-specific terminology and context."""

ARISTOTELOS_SYSTEM_PROMPT = """Communicate only in English.
You construct a semantic network of (composite) concepts around a **[CENTRAL CONCEPT]** based on the user's input.

Follow the seven steps:

1. The user has provided a **[CENTRAL CONCEPT]** and up to 10 related concepts. Use these as a starting point.

2. Extend the set of concepts to 25 by generating additional related concepts that logically fit the network.

3. Provide a detailed definition for the **[CENTRAL CONCEPT]** and short definitions for the related concepts.

4. Describe the relations between each related concept and the **[CENTRAL CONCEPT]**, and between related concepts themselves, using the notation:
   Concept A – relation → Concept B

5. Describe the influence of related concepts on the relations identified in step 4, using the notation:
   Concept A - (Concept C - influence → ) relation → Concept B

6. Turn the relations and influences from steps 4 and 5 into running text. Write a coherent description of the interconnections, structured in paragraphs so that each paragraph covers one logical part.

7. Organise the concepts in a hierarchical structure: place the **[CENTRAL CONCEPT]** at the centre, group related concepts into clusters (e.g. theory, application, critique) and present the structure as a nested list.

Provide only the requested information, without extra explanation. Be structured and complete."""


def load_hierarchy_from_xlsx(path_or_bytes: str | bytes) -> str:
    """
    Leest de kennisclaims-hierarchie uit een xlsx-bestand (bijv. kennisclaims_hierarchie.xlsx).
    Geeft een tekstbeschrijving terug die in de Toulmin-systeemprompt kan worden geïnjecteerd.
    """
    try:
        import openpyxl
    except ImportError:
        return ""
    if isinstance(path_or_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(path_or_bytes, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"### Blad: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_vals = [str(c) if c is not None else "" for c in row]
            if any(row_vals):
                parts.append("\t".join(row_vals))
        parts.append("")
    wb.close()
    return "\n".join(parts).strip()


def run_cimo_toulmin_analysis(
    conjectures_text: str,
    evidence_texts: Optional[List[str]] = None,
    call_llm: Optional[Callable[..., str]] = None,
    hierarchy_text: Optional[str] = None,
    hierarchy_xlsx_path: Optional[str] = None,
    hierarchy_xlsx_bytes: Optional[bytes] = None,
    wisdom_context: Optional[str] = None,
) -> str:
    """
    Runs a CIMO and Toulmin analysis on a document with CIMO conjectures.
    The knowledge level is determined based on indicators and procedures in the documents,
    mapped to the supplied hierarchy (default or from xlsx).
    Optional: wisdom_context = relevant passages from the Cognitia document base (RAG),
    used to determine certainty/knowledge level based on the literature.
    """
    if call_llm is None:
        raise ValueError("call_llm is required")

    # Hiërarchie: expliciet meegegeven tekst, of uit xlsx, anders standaard
    if hierarchy_text:
        hierarchy_block = hierarchy_text
    elif hierarchy_xlsx_path:
        try:
            hierarchy_block = load_hierarchy_from_xlsx(hierarchy_xlsx_path)
            if hierarchy_block:
                hierarchy_block = "## Hiërarchie uit bestand (kennisclaims_hierarchie)\n\n" + hierarchy_block
        except Exception:
            hierarchy_block = HIERARCHY_LEVELS_AND_INDICATORS
    elif hierarchy_xlsx_bytes:
        try:
            hierarchy_block = load_hierarchy_from_xlsx(hierarchy_xlsx_bytes)
            if hierarchy_block:
                hierarchy_block = "## Hiërarchie uit bestand (kennisclaims_hierarchie)\n\n" + hierarchy_block
        except Exception:
            hierarchy_block = HIERARCHY_LEVELS_AND_INDICATORS
    else:
        hierarchy_block = None

    system_prompt = _build_toulmin_system_prompt(hierarchy_block)

    evidence_block = ""
    if evidence_texts:
        evidence_block = "\n\n--- Document(en) met data/bewijs ---\n\n" + "\n\n--- Volgend document ---\n\n".join(evidence_texts)

    wisdom_block = ""
    if wisdom_context and wisdom_context.strip():
        wisdom_block = """

---
## Relevant knowledge from the Cognitia document base (literature)
Use the passages below to assess whether the conjectures are supported or validated in the literature, and to underpin the certainty/knowledge level.

"""
        wisdom_block += wisdom_context.strip()

    user_content = f"""Document with CIMO conjectures:

{conjectures_text}
{evidence_block}{wisdom_block}"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_content},
    ]
    return call_llm(messages, temperature=0.3)


def run_cimo_structuring(
    semantic_network_text: str,
    call_llm: Optional[Callable[..., str]] = None,
) -> str:
    """Structureert een semantisch netwerk volgens CIMO-logica (Denyer)."""
    if call_llm is None:
        raise ValueError("call_llm is required")

    messages = [
        {"role": "system", "content": DENYER_CIMO_SYSTEM_PROMPT},
        {"role": "user", "content": f"""Hier is het semantisch netwerk dat geanalyseerd moet worden:

{semantic_network_text}

Analyseer dit netwerk volgens de CIMO-procedure en geef de conceptuele CIMO-rijen, vermoedens en tabellen."""},
    ]
    return call_llm(messages, temperature=0.3)


def run_semantic_network_builder(
    central_concept: str,
    related_concepts: List[str],
    call_llm: Optional[Callable[..., str]] = None,
) -> str:
    """Bouwt een semantisch netwerk rond een centraal concept (Aristotelos)."""
    if call_llm is None:
        raise ValueError("call_llm is required")

    related_str = ", ".join(related_concepts) if related_concepts else "(geen specifieke concepten gegeven)"
    user_content = f"""CENTRAAL CONCEPT: {central_concept}

Verwante concepten (max 10): {related_str}

Vervaardig het semantisch netwerk volgens de zeven stappen."""

    messages = [
        {"role": "system", "content": ARISTOTELOS_SYSTEM_PROMPT},
        {"role": "user", "content": user_content},
    ]
    return call_llm(messages, temperature=0.3)


EXTRACT_CONCEPTS_SYSTEM_PROMPT = """You analyse a document to extract one central concept and up to 10 related concepts for building a semantic network (Aristoteles).

Give your answer only in the following format, with no extra text before or after:
CENTRAL: <the central main concept in one to a few words>
RELATED: <concept1>, <concept2>, <concept3>, ... (max 10, comma-separated)

The central concept is the main theme or core idea of the document. The related concepts are important related terms that appear in the text. Use only the exact lines CENTRAL: and RELATED:."""


def extract_concepts_from_document(
    doc_text: str,
    call_llm: Optional[Callable[..., str]] = None,
) -> Tuple[str, List[str]]:
    """Extracts central concept and related concepts (max 10) from document text via the LLM."""
    if call_llm is None:
        raise ValueError("call_llm is required")
    if not (doc_text or "").strip():
        return "", []

    # Beperk lengte voor LLM
    text = (doc_text or "").strip()[:12000]

    user_content = f"""Analyse the following document and extract the central concept and up to 10 related concepts.

Document:
{text}

Give your answer in the requested format (CENTRAL: ... and RELATED: ...)."""

    messages = [
        {"role": "system", "content": EXTRACT_CONCEPTS_SYSTEM_PROMPT},
        {"role": "user", "content": user_content},
    ]
    response = call_llm(messages, temperature=0.2).strip()

    central = ""
    related: List[str] = []

    for line in response.split("\n"):
        line = line.strip()
        if line.upper().startswith("CENTRAL:") or line.upper().startswith("CENTRAAL:"):
            central = line.split(":", 1)[-1].strip().strip(".").strip()
        elif line.upper().startswith("RELATED:") or line.upper().startswith("VERWANT:"):
            raw = line.split(":", 1)[-1].strip()
            related = [x.strip() for x in raw.split(",") if x.strip()][:10]

    if not central and related:
        central = related[0] if related else ""
        related = related[1:10]
    return central, related


# Mapping van mogelijke LLM-headers naar Excel-kolomnamen (zoals in technische specificatie)
_TOULMIN_HEADER_MAP = {
    "claim": "CLAIM",
    "data": "DATA",
    "warrant": "WARRANT",
    "backing": "BACKING",
    "rebuttal": "REBUTTAL",
    "qualifier": "QUALIFIER",
    "context (c)": "CONTEXT_C",
    "context": "CONTEXT_C",
    "interventie (i)": "INTERVENTIE_I",
    "interventie": "INTERVENTIE_I",
    "mechanisme (m)": "MECHANISME_M",
    "mechanisme": "MECHANISME_M",
    "outcome (o)": "OUTCOME_O",
    "outcome": "OUTCOME_O",
    "gevon indicatoren": "INDICATOR_VECTOR",  # typo-variant
    "gevonden indicatoren": "INDICATOR_VECTOR",
    "verwijzingen naar procedures": "PROCEDURE_VECTOR",
    "level_nummer": "LEVEL_NUMBER",
    "level nummer": "LEVEL_NUMBER",
    "level_label": "LEVEL_LABEL",
    "level label": "LEVEL_LABEL",
    "zekerheidsniveau": "CERTAINTY_SCORE",
    "explicietheid": "EXPLICIETHEID_SCORE",
    "kennisniveau": "LEVEL",
    "onderbouwing": "EVIDENCE_MATRIX",
    "cimo-analyse": "CIMO_ANALYSE",
}
# Volgorde kolommen in Excel (technische specificatie)
_TOULMIN_EXCEL_COLUMNS = [
    "CLAIM", "DATA", "WARRANT", "BACKING", "REBUTTAL", "QUALIFIER",
    "LEVEL", "INDICATOR_VECTOR", "PROCEDURE_VECTOR", "CERTAINTY_SCORE",
    "EXPLICIETHEID_SCORE", "LEVEL_NUMBER", "LEVEL_LABEL", "EVIDENCE_MATRIX",
    "CONTEXT_C", "INTERVENTIE_I", "MECHANISME_M", "OUTCOME_O", "CIMO_ANALYSE",
]


def _normalize_header(s: str) -> Optional[str]:
    """Herkent een regel als header en geeft de genormaliseerde Excel-kolomnaam terug."""
    s = s.strip()
    # Alleen regels die met **, -, of een bekend label gevolgd door : beginnen
    if not (
        s.startswith(("*", "-"))
        or re.match(
            r"^(claim|data|warrant|backing|rebuttal|qualifier|level[_ ]?(nummer|label)?|"
            r"gevonden indicatoren|verwijzingen naar procedures|zekerheidsniveau|"
            r"explicietheid|kennisniveau|onderbouwing|context|interventie|mechanisme|outcome)\s*:",
            s,
            re.IGNORECASE,
        )
    ):
        return None
    # **Claim**: of **Claim** of - **Claim**: of Claim:
    m = re.match(r"^[-*]*\s*\*{0,2}([^*:]+)\*{0,2}\s*:?\s*(.*)$", s, re.IGNORECASE)
    if m:
        label = m.group(1).strip().lower()
        for key, col in _TOULMIN_HEADER_MAP.items():
            if key in label or label in key:
                return col
        if "claim" in label:
            return "CLAIM"
        if "data" in label:
            return "DATA"
        if "warrant" in label:
            return "WARRANT"
        if "qualifier" in label:
            return "QUALIFIER"
        if "level" in label and "nummer" in label:
            return "LEVEL_NUMBER"
        if "level" in label and "label" in label:
            return "LEVEL_LABEL"
        if "indicator" in label or "indicatoren" in label:
            return "INDICATOR_VECTOR"
        if "procedure" in label:
            return "PROCEDURE_VECTOR"
        if "zekerheid" in label:
            return "CERTAINTY_SCORE"
        if "explicietheid" in label:
            return "EXPLICIETHEID_SCORE"
        if "onderbouwing" in label:
            return "EVIDENCE_MATRIX"
        if "context" in label:
            return "CONTEXT_C"
        if "interventie" in label:
            return "INTERVENTIE_I"
        if "mechanisme" in label:
            return "MECHANISME_M"
        if "outcome" in label:
            return "OUTCOME_O"
    # Alleen "Header:" zonder sterretjes
    for key, col in _TOULMIN_HEADER_MAP.items():
        if s.lower().startswith(key + ":"):
            return col
    if re.match(r"^[a-z\s]+:\s*", s.lower()):
        pre = s.split(":", 1)[0].strip().lower()
        for k, col in _TOULMIN_HEADER_MAP.items():
            if k in pre or pre in k:
                return col
    return None


def parse_toulmin_results_for_excel(analysis_text: str) -> List[dict]:
    """
    Parset de LLM-tekst naar gestructureerde rijen voor Excel.
    Herkent headers zoals **Claim**:, **Data**:, Level_nummer:, etc. en vult alle kolommen.
    Eén rij per 'vermoeden'; bij nieuwe **Claim** start een nieuwe rij.
    """
    rows: List[dict] = []
    current: dict = {}
    last_key: Optional[str] = None
    lines = analysis_text.split("\n")

    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        col = _normalize_header(line_stripped)
        if col is not None:
            # Nieuwe header: bij **Claim** of LEVEL_NUMBER e.d. een nieuwe rij starten als we al een Claim hebben
            if col == "CLAIM" and current and current.get("CLAIM"):
                rows.append(current)
                current = {}
            # Waarde: tekst na de eerste dubbele punt
            if ":" in line_stripped:
                value = line_stripped.split(":", 1)[-1].strip().strip("-* ")
            else:
                value = ""
            current[col] = value
            last_key = col
            continue
        # Geen header: voortzetting van vorige veld
        if last_key and current is not None:
            current[last_key] = (current.get(last_key) or "") + "\n" + line_stripped

    if current:
        rows.append(current)

    if not rows:
        rows = [{"CLAIM": analysis_text[:32000]}]

    # Zorg dat elke rij alle kolommen heeft (leeg waar niet gevuld)
    out = []
    for r in rows:
        row = {}
        for c in _TOULMIN_EXCEL_COLUMNS:
            row[c] = r.get(c, "")
        for k, v in r.items():
            if k not in row:
                row[k] = v
        out.append(row)
    return out


def _clean_excel_text(val: str) -> str:
    """
    Remove characters that are illegal in Excel's XML representation.
    This avoids openpyxl IllegalCharacterError when writing LLM output.
    """
    if not isinstance(val, str):
        return val
    # Excel (XML 1.0) does not allow control chars except tab, LF, CR.
    # Replace disallowed control characters with a space.
    return "".join(
        ch if ch in ("\t", "\n", "\r") or ord(ch) >= 32 else " "
        for ch in val
    )


def build_toulmin_excel(analysis_text: str) -> bytes:
    """
    Maakt een Excel-bestand van de Toulmin/CIMO-analyse.
    Gebruikt openpyxl indien beschikbaar, anders een CSV in memory (als .csv string).
    """
    rows = parse_toulmin_results_for_excel(analysis_text)

    if HAS_OPENPYXL and rows:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CIMO-Toulmin analyse"
        header_font = Font(bold=True)
        all_keys = set()
        for r in rows:
            all_keys.update(r.keys())
        # Vaste kolomvolgorde (technische specificatie); daarna eventuele extra keys
        headers = list(_TOULMIN_EXCEL_COLUMNS)
        for k in sorted(all_keys):
            if k not in headers:
                headers.append(k)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = header_font
        for row_idx, row in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                if isinstance(val, str):
                    # Truncate very long strings for Excel cell limit
                    if len(val) > 32767:
                        val = val[:32700] + "..."
                    # Strip illegal XML characters
                    val = _clean_excel_text(val)
                ws.cell(row=row_idx, column=col_idx, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # Geen openpyxl of lege rijen: CSV
    import csv
    buf = io.StringIO()
    if rows:
        all_keys = list(rows[0].keys())
        writer = csv.DictWriter(buf, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(rows)
    else:
        buf.write("Analyse\n")
        buf.write(analysis_text.replace("\n", " "))
    return buf.getvalue().encode("utf-8-sig")


def get_excel_extension() -> str:
    return ".xlsx" if HAS_OPENPYXL else ".csv"


def build_toulmin_json(analysis_text: str) -> str:
    """
    Exporteert de Toulmin-analyse als JSON (technische specificatie: evidence_matrix / results).
    Bevat per claim: claim, data, warrant, qualifier, level_number, level_label,
    zekerheidsniveau, explicietheid, gevonden indicatoren, procedures.
    """
    rows = parse_toulmin_results_for_excel(analysis_text)
    # Normaliseer keys voor JSON (behoud alle velden)
    out = {
        "source": "GDGO_Toulmin_Weaver",
        "hierarchy_reference": "2025 11 13 kennisclaims_hierarchie",
        "claims": [],
    }
    for i, row in enumerate(rows):
        claim_entry = {"index": i + 1}
        for k, v in row.items():
            if v and isinstance(v, str) and len(v) > 5000:
                v = v[:5000] + "..."
            claim_entry[k] = v
        out["claims"].append(claim_entry)
    return json.dumps(out, ensure_ascii=False, indent=2)
